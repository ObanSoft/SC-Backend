from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from app.models.Producto import Producto

def exportar_inventario_servicio():
    try:
        productos = Producto.query.filter_by(estado='inventario').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        encabezados = ['ID', 'Código', 'Nombre', 'Precio', 'Fecha de Creación']
        ws.append(encabezados)

        for p in productos:
            ws.append([
                p.id,
                p.identificador_unico,
                p.nombre,
                float(p.precio),
                p.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if p.fecha_creacion else ''
            ])

        bold_font = Font(bold=True, color='FFFFFF')
        fill_color = PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )

        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = fill_color
            cell.alignment = center_align
            cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
            for i, cell in enumerate(row):
                cell.border = border
                cell.alignment = Alignment(horizontal='left')
                if i == 3:  # columna precio
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE.replace('$', '$')

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        table_range = f"A1:E{ws.max_row}"
        tabla = Table(displayName="TablaInventario", ref=table_range)
        estilo_tabla = TableStyleInfo(
            name="TableStyleMedium10",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo_tabla
        ws.add_table(tabla)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output, None
    except Exception as e:
        return None, str(e)
