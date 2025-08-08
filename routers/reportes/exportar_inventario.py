# routers/productos/exportar_inventario.py
from flask import Blueprint, request, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from models.Producto import Producto
from utils.auth_utils import token_required
from flask_cors import cross_origin

exportar_inventario_bp = Blueprint('exportar_inventario', __name__)

@exportar_inventario_bp.route('/exportar_inventario', methods=['GET', 'OPTIONS'], strict_slashes=False)
@cross_origin(origin='http://localhost:3000', supports_credentials=True)
@token_required
def descargar_excel_inventario():
    if request.method == 'OPTIONS':
        return '', 200

    productos = Producto.query.filter_by(estado='inventario').all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    encabezados = ['ID', 'Código', 'Nombre', 'Precio', 'Fecha de Creación']
    ws.append(encabezados)

    for p in productos:
        ws.append([
            p.id,
            p.identificador_unico,
            p.nombre,
            float(p.precio),
            p.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if p.fecha_creacion else ''
        ])

    # Estilos
    bold_font = Font(bold=True, color='FFFFFF')
    fill_color = PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = fill_color
        cell.alignment = center_align
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for i, cell in enumerate(row):
            cell.border = border
            cell.alignment = Alignment(horizontal='left')
            if i == 3:  # columna precio
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE.replace('$', '$')

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    table_range = f"A1:E{ws.max_row}"
    tabla = Table(displayName="TablaInventario", ref=table_range)
    estilo_tabla = TableStyleInfo(name="TableStyleMedium10", showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_inventario.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )