from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from app.models.Venta import Venta

def exportar_ventas_servicio():
    try:
        ventas = Venta.query.filter_by(tipo_venta='Individual').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas Individuales"

        encabezados = [
            'ID', 
            'Identificador Único', 
            'Nombre Producto', 
            'Precio', 
            'Fecha de Venta',
            'Vendido Por',
            'Método de Pago'
        ]
        ws.append(encabezados)

        for venta in ventas:
            ws.append([
                venta.id,
                venta.identificador_unico,
                venta.nombre_producto,
                float(venta.precio),
                venta.fecha_venta.strftime('%Y-%m-%d %H:%M:%S') if venta.fecha_venta else '',
                venta.vendido_por or '',
                venta.metodo_pago or ''
            ])

        # Estilos
        bold_font = Font(bold=True, color='FFFFFF')
        fill_color = PatternFill(start_color='E91E63', end_color='E91E63', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )

        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = fill_color
            cell.alignment = center_align
            cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
            for idx, cell in enumerate(row, start=1):
                cell.border = border
                cell.alignment = Alignment(horizontal='left')
                if idx == 4:  # columna precio
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE.replace('$', '$')

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        table_range = f"A1:G{ws.max_row}"
        tabla = Table(displayName="TablaVentas", ref=table_range)
        estilo_tabla = TableStyleInfo(
            name="TableStyleMedium10",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo_tabla
        ws.add_table(tabla)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output, None
    except Exception as e:
        return None, str(e)
